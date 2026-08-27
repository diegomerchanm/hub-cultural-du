"""
load_manual_account_categorization.py — Hub Cultural DU

Sube la categorización manual curada (arte, tipo de institución, identidad
cultural, zona geográfica, seguidores verificados, etc.) desde el Excel de
trabajo hacia los nodos :Account en Neo4j, usando MERGE por username:

  - Si la cuenta ya existe (fue harvesteada por las fases 1-2), le añade
    estas propiedades sin tocar followersCount, bio, culturalRelevanceScore
    ni ninguna otra propiedad que ya gestiona el pipeline automático.
  - Si la cuenta todavía no fue harvesteada, crea el nodo :Account de una
    vez (sin labels :Public/:Private — esos los pone 2_build_graph.py
    cuando corra la extracción real).

Todas las propiedades usan nombres propios (artType, institutionType,
manualFollowersCount, ...) que no colisionan con lo que escribe
2_build_graph.py / run_gds_algorithms.py — así que reprocesar las fases 1-2
después de correr este script no pisa esta data curada a mano, y viceversa.

Idempotente: se puede volver a correr tras cada actualización del Excel
(MERGE + SET siempre deja el nodo con el valor más reciente de la hoja).

Uso:
    python load_manual_account_categorization.py --dry-run
    python load_manual_account_categorization.py
"""

import os
import re
from datetime import date

import openpyxl
import typer
from dotenv import load_dotenv
from neo4j import GraphDatabase

load_dotenv()
NEO4J_URI = os.getenv("NEO4J_URI")
NEO4J_USERNAME = os.getenv("NEO4J_USERNAME")
NEO4J_PASSWORD = os.getenv("NEO4J_PASSWORD")

app = typer.Typer()

DEFAULT_XLSX = "cuentas_instagram_completo_v4.xlsx"

# columna Excel (1-indexed) -> propiedad Neo4j
COLUMN_MAP = {
    2: "artType",                    # Tipo de arte
    3: "institutionType",            # Tipo de institución
    4: "eventFrequency",             # Frecuencia de eventos
    5: "parentInstitution",          # Es parte de institución mayor (+ link)
    6: "contentType",                # Tipo de contenido
    8: "hasFreeEvents",              # Tiene eventos gratis
    9: "priceRange",                 # Rango de precio
    10: "promotedOutsideInstagram",  # Eventos fuera de Instagram (+ fuente)
    11: "eventFormat",               # Curso / evento único / requisitos
    12: "culturalIdentity",          # Identidad cultural (si aplica)
    13: "geoZone",                   # Île-de-France / Francia fuera IDF / Fuera de Francia
    14: "photoPermissionRaw",        # Permiso para mostrar sus fotos en el sitio (Sí/No, vacío = sin dato)
}


def _is_yes(raw) -> bool | None:
    """Normaliza la columna 14 (texto libre en la planilla, ej. 'Sí',
    'si', 'Sí, ya la contactamos') a True/False/None. None = sin dato
    todavía (cuenta sin contactar aún) — en el sitio se trata igual que
    False (default seguro: sin permiso explícito, no se muestra la foto
    real), pero se guarda distinto para poder filtrar después quién
    falta contactar."""
    if not isinstance(raw, str) or not raw.strip():
        return None
    return raw.strip().lower().startswith(("s", "y"))  # sí / si / yes


def handle_rows():
    """Reconstruye la lista de filas que tienen el handle real (columna A),
    saltando las filas de descripción intercaladas que quedaron del Excel
    original garabateado: cuenta N -> fila 2N para N<=38, fila 78 para N=39
    (cuenta huérfana sin fila de descripción pareada), fila 2N-1 para N>=40
    hasta la fila 251 (cuenta 126, la última del lote original curado a mano).

    Filas 253-413 (2026-08-21): 161 cuentas nuevas (bucket "fijo" de
    pilot_classification.csv) categorizadas vía subagentes Haiku con
    búsqueda web — ver decisions_es.md. A diferencia del bloque anterior,
    estas NO llevan fila de descripción intercalada (una fila por cuenta,
    contigua) porque el script nunca leyó esas filas de todos modos — son
    puramente cosméticas para navegación visual en Excel. Filas 282 y 332
    quedaron vacías a propósito (username en columna A = None, `load_rows`
    las salta): `semaine_de_la_critique` y `citedelabd` ya estaban en el
    bloque original (filas 95 y 109) con datos más completos — se detectó
    el duplicado por username y se priorizó la entrada original.
    """
    rows = [2 * n for n in range(1, 39)]
    rows.append(78)
    rows += [2 * n - 1 for n in range(40, 127)]
    rows += list(range(253, 414))
    return rows


def parse_followers(raw):
    """'132K seguidores' -> 132000, '51.4K seguidores' -> 51400,
    '~33.000 (IG)' -> 33000, '4,377 seguidores' -> 4377, '70 (IG...)' -> 70.
    Devuelve None si no hay ningún número reconocible."""
    if not raw:
        return None
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*([KkMm])\b", raw)
    if m:
        base = float(m.group(1).replace(",", "."))
        mult = 1_000 if m.group(2).upper() == "K" else 1_000_000
        return round(base * mult)
    m = re.search(r"(\d[\d.,]*)", raw)
    if m:
        digits = re.sub(r"[.,]", "", m.group(1))
        return int(digits) if digits else None
    return None


def load_rows(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Hoja1"]
    out = []
    for r in handle_rows():
        username = (ws.cell(row=r, column=1).value or "").strip()
        if not username:
            continue
        props = {"username": username}
        for col, key in COLUMN_MAP.items():
            val = ws.cell(row=r, column=col).value
            props[key] = val.strip() if isinstance(val, str) else val
        raw_followers = ws.cell(row=r, column=7).value
        props["manualFollowersLabel"] = raw_followers
        props["manualFollowersCount"] = parse_followers(raw_followers or "")
        props["photoPermission"] = _is_yes(props.pop("photoPermissionRaw", None))
        out.append(props)
    return out


UPSERT_QUERY = """
UNWIND $rows AS row
MERGE (a:Account {username: row.username})
ON CREATE SET a.firstSeenAt = datetime()
SET a.artType                = row.artType,
    a.institutionType         = row.institutionType,
    a.eventFrequency          = row.eventFrequency,
    a.parentInstitution       = row.parentInstitution,
    a.contentType             = row.contentType,
    a.hasFreeEvents           = row.hasFreeEvents,
    a.priceRange              = row.priceRange,
    a.promotedOutsideInstagram = row.promotedOutsideInstagram,
    a.eventFormat             = row.eventFormat,
    a.culturalIdentity        = row.culturalIdentity,
    a.geoZone                 = row.geoZone,
    a.manualFollowersLabel    = row.manualFollowersLabel,
    a.manualFollowersCount    = row.manualFollowersCount,
    a.photoPermission         = row.photoPermission,
    a.manualDataCuratedAt     = $curatedAt,
    a.manualDataSource        = $source
RETURN count(a) AS n
"""


@app.command()
def main(
    xlsx: str = typer.Option(DEFAULT_XLSX, help="Ruta al Excel curado"),
    dry_run: bool = typer.Option(False, "--dry-run", help="No escribe en Neo4j, solo muestra un resumen"),
):
    rows = load_rows(xlsx)
    print(f"📄 {len(rows)} cuentas leídas de {xlsx}")

    missing_geo = [r["username"] for r in rows if not r.get("geoZone")]
    missing_followers = [r["username"] for r in rows if r.get("manualFollowersCount") is None]
    if missing_geo:
        print(f"⚠️  Sin geoZone: {missing_geo}")
    if missing_followers:
        print(f"⚠️  Sin follower count parseable: {missing_followers}")

    if dry_run:
        print("\n[dry-run] Ejemplo de las primeras 3 filas a escribir:")
        for r in rows[:3]:
            print(" ", r)
        print("\n[dry-run] No se escribió nada en Neo4j.")
        return

    if not all([NEO4J_URI, NEO4J_USERNAME, NEO4J_PASSWORD]):
        raise ValueError("Error: credenciales Neo4j ausentes en .env")

    driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USERNAME, NEO4J_PASSWORD))
    driver.verify_connectivity()

    curated_at = date.today().isoformat()
    source = "Curacion manual (WebSearch + verificacion Instagram via Claude in Chrome), ago 2026"

    with driver.session() as session:
        result = session.run(UPSERT_QUERY, rows=rows, curatedAt=curated_at, source=source)
        n = result.single()["n"]
        print(f"\n✅ {n} nodos :Account actualizados/creados con categorizacion manual.")

    driver.close()


if __name__ == "__main__":
    app()
